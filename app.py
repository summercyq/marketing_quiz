import streamlit as st
import pandas as pd
import random
import os
from datetime import datetime
from openpyxl import load_workbook # 載入openpyxl用於寫入Excel (管理者功能)

st.set_page_config(page_title="TIMS行銷專業能力認證 2025(初級)題庫", layout="wide")
st.title("TIMS行銷專業能力認證 2025(初級)題庫")

# 檔案路徑設定
EXCEL_PATH = "行銷題庫總表.xlsx"
SHEET_NAME = "題庫總表"
WRONG_LOG = "錯題紀錄.csv"
STATS_LOG = "答題統計.csv" # 答題統計功能未在原碼中實現，但路徑已定義
EDIT_PASSWORD = "quiz2024" # 管理者密碼

# 使用st.cache_data載入資料，避免每次重跑都重新載入
@st.cache_data
def load_data():
    """Loads the question data from the Excel file."""
    try:
        # 嘗試讀取 Excel 檔案，如果不存在或有問題，會捕獲異常
        if not os.path.exists(EXCEL_PATH):
             st.error(f"錯誤：找不到題庫檔案 `{EXCEL_PATH}`。請確認檔案是否存在。")
             return pd.DataFrame() # Return empty dataframe if file not found

        # 檢查檔案是否為空的
        if os.path.getsize(EXCEL_PATH) == 0:
             st.error(f"錯誤：題庫檔案 {EXCEL_PATH} 為空。請確認檔案內容。")
             return pd.DataFrame()

        # 檢查工作表是否存在
        try:
             wb = load_workbook(EXCEL_PATH, read_only=True)
             if SHEET_NAME not in wb.sheetnames:
                  st.error(f"錯誤：題庫檔案 {EXCEL_PATH} 中找不到工作表 `{SHEET_NAME}`。")
                  return pd.DataFrame()
        except Exception as wb_e:
             st.error(f"無法打開或檢查題庫檔案 {EXCEL_PATH} 的工作表：{wb_e}")
             return pd.DataFrame()

        # 讀取工作表
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)

        # 檢查 DataFrame 是否為空 (可能工作表存在但內容是空的)
        if df.empty:
             st.error(f"錯誤：題庫檔案 {EXCEL_PATH} 的 {SHEET_NAME} 工作表內容為空。")
             return pd.DataFrame()

        # 簡易檢查必要欄位是否存在 (至少要有 章節, 題號, 題目, 解答)
        required_cols = ["章節", "題號", "題目", "解答"]
        if not all(col in df.columns for col in required_cols):
             missing = [col for col in required_cols if col not in df.columns]
             st.error(f"錯誤：題庫檔案 {EXCEL_PATH} 的 {SHEET_NAME} 工作表缺少必要欄位：{', '.join(missing)}。")
             return pd.DataFrame()


        return df # 返回成功載入的 DataFrame

    except FileNotFoundError:
        # This should ideally be caught by the initial os.path.exists check, but kept for robustness
        st.error(f"錯誤：找不到題庫檔案 `{EXCEL_PATH}`。請確認檔案是否存在。")
        return pd.DataFrame() # Return empty dataframe on error
    except Exception as e:
        # Catch any other reading errors (e.g., corrupted file, incorrect format)
        st.error(f"載入題庫時發生錯誤：{e}")
        return pd.DataFrame()

df = load_data()


# 章節對應關係 (CH10 已移除)
# 使用字典 comprehension 建立章節對應，更簡潔
chapter_mapping = {f"CH{i}": [f"{i}-1", f"{i}-2"] for i in range(1, 10)}
# 驗證載入的資料是否有對應的章節/子章節 (可選的健壯性檢查)
if not df.empty:
    # 獲取所有在 chapter_mapping 中定義的子章節
    all_defined_sections = [sec for ch_sections in chapter_mapping.values() for sec in ch_sections]
    # 過濾掉 DataFrame 中不在定義範圍內的章節 (原始資料可能包含舊的或錯誤的章節)
    df = df[df["章節"].astype(str).isin(all_defined_sections)].copy() # 使用 .copy() 避免 SettingWithCopyWarning
    if df.empty and not load_data().empty: # 如果原始檔案非空但過濾後為空
         st.warning("題庫檔案已載入，但找不到符合定義章節範圍 (CH1-CH9, 每個章節含 -1, -2 小節) 的題目。")


# 初始化 Session State
# 為每個需要在不同執行之間保留狀態的變數設定預設值
for key in ["quiz_started", "questions", "user_answers", "shuffled_options", "last_settings", "is_admin_mode", "username"]:
    # 如果 Session State 中沒有這個 key，則設定其預設值
    if key not in st.session_state:
        if key == "is_admin_mode" or key == "quiz_started":
            st.session_state[key] = False # 預設不啟動測驗或管理者模式
        elif key == "user_answers":
             st.session_state[key] = [] # 用於存放使用者回答的列表
        elif key == "questions":
             st.session_state[key] = pd.DataFrame() # 用於存放本次測驗題目的 DataFrame
        elif key == "shuffled_options":
             st.session_state[key] = {} # 用於存放每道題目的選項打亂順序
        elif key == "last_settings":
             st.session_state[key] = None # 用於存放上一次測驗設定，方便重新開始
        elif key == "username":
             st.session_state[key] = "" # 用於存放使用者名稱

# --- Helper function to generate quiz questions ---
def generate_quiz_questions(username, mode, selected_chapters, num_questions, dataframe, chapter_map, wrong_log_path):
    """Generates a list of questions based on the selected mode and settings."""
    if dataframe.empty:
         # Warning already shown by load_data or filtering, just return empty
         return pd.DataFrame()

    filtered = pd.DataFrame() # Initialize empty filtered dataframe

    if mode == "一般出題模式":
        if not selected_chapters:
             st.warning("請至少選擇一個章節。")
             return pd.DataFrame()
        # Map selected chapters (e.g., CH1) to sections (e.g., 1-1, 1-2)
        sections = [s for ch in selected_chapters for s in chapter_map.get(ch, [])]
        if not sections: # Should not happen if chapter_map is correct and selected_chapters are valid keys
             st.warning(f"所選章節 ({', '.join(selected_chapters)}) 無法對應到任何子章節。")
             return pd.DataFrame()
        # Filter dataframe by the sections
        filtered = dataframe[dataframe["章節"].astype(str).isin(sections)].copy() # Use .copy()

        if filtered.empty:
             st.warning(f"找不到符合所選章節 ({', '.join(selected_chapters)}) 的題目。請檢查題庫檔案或章節選擇。")
             return pd.DataFrame()

    elif mode == "錯題再練模式":
        if not username.strip():
             st.warning("錯題再練模式需要使用者名稱。請輸入使用者名稱。")
             return pd.DataFrame()

        if os.path.exists(wrong_log_path):
            try:
                log = pd.read_csv(wrong_log_path)
                # Filter log for the current user (case-insensitive)
                user_wrong_log = log[log["使用者"].str.lower() == username.lower()].copy() # Use .copy()

                # Apply chapter filter if selected_chapters is not empty
                if selected_chapters:
                    sections = [s for ch in selected_chapters for s in chapter_map.get(ch, [])]
                    user_wrong_log = user_wrong_log[user_wrong_log["章節"].astype(str).isin(sections)].copy() # Use .copy()

                if user_wrong_log.empty:
                     if selected_chapters:
                         st.info(f"使用者 {username} 在所選章節 ({', '.join(selected_chapters)}) 中沒有錯題紀錄。")
                     else:
                         st.info(f"使用者 {username} 沒有錯題紀錄。")
                     return pd.DataFrame()

                # Merge with the main dataframe to get full question details for the wrong questions
                # Use drop_duplicates in case a question is in the log multiple times
                wrong_question_keys = user_wrong_log[["章節", "題號"]].drop_duplicates()
                # Ensure keys are comparable types, e.g., strings
                wrong_question_keys["章節"] = wrong_question_keys["章節"].astype(str)
                wrong_question_keys["題號"] = wrong_question_keys["題號"].astype(str)
                dataframe["章節"] = dataframe["章節"].astype(str)
                dataframe["題號"] = dataframe["題號"].astype(str)

                # Merge based on chapter and question number
                filtered = dataframe.merge(
                    wrong_question_keys,
                    on=["章節", "題號"]
                ).copy() # Use .copy()

                if filtered.empty:
                     st.warning(f"根據使用者 {username} 的錯題紀錄，在題庫中找不到對應的題目。請確認題庫檔案。")
                     return pd.DataFrame() # Should not happen if user_wrong_log is not empty and merge keys are correct

            except pd.errors.EmptyDataError:
                 st.info("錯題紀錄檔案為空。")
                 return pd.DataFrame()
            except FileNotFoundError:
                 st.info("找不到錯題紀錄檔案。請先進行作答以產生紀錄。")
                 return pd.DataFrame()
            except Exception as e:
                 st.error(f"讀取錯題紀錄時發生錯誤：{e}")
                 return pd.DataFrame()

        else:
            st.info("找不到錯題紀錄檔案。請先進行作答以產生紀錄。")
            return pd.DataFrame()

    else: # Should not happen with the new structure
        st.error("內部錯誤：無效的測驗模式選擇。")
        return pd.DataFrame()

    # Sample questions
    if not filtered.empty:
        # Ensure num_questions does not exceed the available questions
        actual_num_questions = min(num_questions, len(filtered))
        if actual_num_questions == 0:
             st.warning("找不到符合條件的題目，無法出題。")
             return pd.DataFrame()
        return filtered.sample(n=actual_num_questions, random_state=None).reset_index(drop=True).copy() # Use .copy() and random_state=None for different samples

    else:
        # This case is handled by the empty checks within each mode block
        return pd.DataFrame()


# --- Sidebar ---
st.sidebar.header("使用者與模式設定")
st.session_state.username = st.sidebar.text_input("請輸入使用者名稱", value=st.session_state.get("username", ""), key="username_input")


# --- Sidebar - Quiz Settings (Only display if not in admin mode) ---
# The admin mode checkbox is now placed BELOW the start quiz button in the sidebar.
# Initialize the checkbox state early based on session_state
is_admin_mode_initial = st.session_state.get("is_admin_mode", False)
st.session_state.is_admin_mode = st.sidebar.checkbox("🛠️ 啟用管理者模式", value=is_admin_mode_initial, key="admin_mode_checkbox")


if not st.session_state.is_admin_mode:
    # These inputs only show if not in admin mode
    quiz_mode = st.sidebar.radio("選擇模式：", ["一般出題模式", "錯題再練模式"], key="quiz_mode_radio")
    selected_chapters = st.sidebar.multiselect("選擇章節：", list(chapter_mapping.keys()), default=st.session_state.last_settings.get("selected_chapters", ["CH1"]) if st.session_state.last_settings else ["CH1"], key="chapters_select")
    num_questions = st.sidebar.number_input("出題數量", min_value=1, max_value=50, value=st.session_state.last_settings.get("num_questions", 5) if st.session_state.last_settings else 5, key="num_questions_input")

    # Start Quiz Button
    if st.sidebar.button("🚀 開始出題", key="start_quiz_button"):
        if not st.session_state.username.strip():
            st.sidebar.warning("請先輸入使用者名稱！")
        elif df.empty:
             st.sidebar.warning("題庫資料為空，無法開始測驗。請通知管理者檢查題庫檔案。")
        else:
